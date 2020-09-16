# _*_ coding: utf-8 _*_
# @Time    :2020/7/2
# @Author  :lemon_yuehaulong
# @Email    :374748859@qq.com
#例
from openpyxl import load_workbook
from openpyxl import  Workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]
    for i in range(2,sheet.max_row+1):
        case=[]
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i,column=j).value)
        all_case.append(case)
    print(all_case)
    return all_case
if __name__ == '__main__':      #只在本页执行此打印main
    read_data('text_shuju_1.xlsx','Sheet1')


