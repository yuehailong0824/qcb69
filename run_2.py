
from read_data import read_data
from request import http_request
from openpyxl import load_workbook
Token=None
def run_2(file_name,sheet_name,c1,c2):
    global Token
    all_case=read_data(file_name,sheet_name)
    ip = "http://120.78.128.25:8766"
    for test_data in all_case:
        response = http_request(ip + test_data[4], eval(test_data[5]), token=Token, method=test_data[3])
        if 'login' in test_data[4]:
            Token = "Bearer "+response['data']['token_info']['token']
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(row=test_data[0]+1,column=c1).value=str(response)
        actual={'code':response['code'],'msg':response['msg']}
        if eval(test_data[6])==actual:
            print('测试通过')
            sheet.cell(row=test_data[0]+1,column=c2).value='PASS'
        else:
            print('测试不通过')
            sheet.cell(row=test_data[0]+1,column=c2).value='FAIL'
        wb.save('text_shuju_1.xlsx')
run_2('text_shuju_1.xlsx','Sheet1',8,9)



